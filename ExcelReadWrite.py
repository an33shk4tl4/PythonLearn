from openpyxl import Workbook, cell
wb = Workbook()

wb.create_sheet("Halloweensht")
# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
# ws['A2'] = datetime.datetime.now()
ws.append([datetime.datetime.now()])

# TODO - figure out mutli cell editing some day.. :)
# ws['C1:D2'] = ((100, 222), (888, 3333))

# Save the file
wb.save("sample.xlsx")
